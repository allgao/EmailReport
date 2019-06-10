import os
import smtplib
import openpyxl
import jinja2
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader


class Performance:
    def __init__(self, caseCount, amount, paybackAmount, paybackRate, targetRate, completionRate, completionAmount, order, _15daysPayback, _20daysPayback, _15daysCompletionRate, _20daysCompletionRate):
        self.caseCount = '{:,}'.format(caseCount)
        self.amount = '{:,}'.format(amount)
        self.paybackAmount = '{:,.0f}'.format(paybackAmount)
        self.paybackRate = '{:.2%}'.format(paybackRate)
        self.targetRate = '{:.2%}'.format(targetRate)
        self.completionRate = '{:.2%}'.format(completionRate)
        self.completionAmount = '{:,.0f}'.format(completionAmount)
        self.order = order
        self._15daysPayback = '{:,.0f}'.format(_15daysPayback)
        self._20daysPayback = '{:,.0f}'.format(_20daysPayback)
        self._15daysCompletionRate = '{:.2%}'.format(_15daysCompletionRate)
        self._20daysCompletionRate = '{:.2%}'.format(_20daysCompletionRate)


class OverallPerformance:
    def __init__(self, caseCount, amount, paybackAmount, paybackRate, targetOne, targetTwo, targetThree, targetThreeCompletionRate, targetOnePaybackAmount, targetTwoPaybackAmount, targetThreePaybackAmount, order):
        self.caseCount = '{:,}'.format(caseCount)
        self.amount = '{:,.0f}'.format(amount)
        self.paybackAmount = '{:,.0f}'.format(paybackAmount)
        self.paybackRate = '{:.2%}'.format(paybackRate)
        self.targetOne = '{:.2%}'.format(targetOne)
        self.targetTwo = '{:.2%}'.format(targetTwo)
        self.targetThree = '{:.2%}'.format(targetThree)
        self.targetThreeCompletionRate = '{:.2%}'.format(
            targetThreeCompletionRate)
        self.targetOnePaybackAmount = '{:,.0f}'.format(targetOnePaybackAmount)
        self.targetTwoPaybackAmount = '{:,.0f}'.format(targetTwoPaybackAmount)
        self.targetThreePaybackAmount = '{:,.0f}'.format(
            targetThreePaybackAmount)
        self.order = order


class PerformanceReport:
    def __init__(self, companyName, _180Performance, _180GreaterPerformance, overallPerformance):
        self.companyName = companyName
        self._180Performance = _180Performance
        self._180GreaterPerformance = _180GreaterPerformance
        self.overallPerformance = overallPerformance


class Contact:
    def __init__(self, receivers, ccs):
        self.receivers = receivers
        self.ccs = ccs


def loadContacts():
    wb = load_workbook("业绩模板.xlsx", data_only=True)
    ws = wb[wb.sheetnames[1]]
    d = {}
    for i in range(2, ws.max_row+1):
        receivers = ws.cell(row=i, column=2).value
        ccs = ws.cell(row=i, column=3).value
        contact = Contact(receivers, ccs)
        d[ws.cell(row=i, column=1).value] = contact
    return d


def generateEmailAndSend(companyName, d, htmlContent, smtp):
    if companyName in d.keys():
        msg = MIMEMultipart()
        msg['Subject'] = companyName + "业绩统计"
        msg['From'] = SENDER
        msg['To'] = d[companyName].receivers
        msg['Cc'] = d[companyName].ccs
        content = MIMEText(htmlContent, 'html')
        msg.attach(content)
        smtp.sendmail(SENDER, d[companyName].receivers, msg.as_string())
        print(companyName+" 邮件发送成功!")
    else:
        print(companyName+"邮件发送失败！ 找不到联系人信息")


def loadRowData():
    wb = load_workbook("业绩模板.xlsx", data_only=True)
    ws = wb[wb.sheetnames[0]]
    dataList = []
    for x in range(3, 1000):
        if "总计" == ws.cell(row=x, column=1).value:
            break
        else:
            _180Performance = Performance(caseCount=ws.cell(row=x, column=2).value, amount=ws.cell(row=x, column=3).value, paybackAmount=ws.cell(row=x, column=4).value, paybackRate=ws.cell(row=x, column=5).value, targetRate=ws.cell(row=x, column=6).value, completionRate=ws.cell(row=x, column=7).value, completionAmount=ws.cell(
                row=x, column=8).value, order=ws.cell(row=x, column=9).value, _15daysPayback=ws.cell(row=x, column=10).value, _20daysPayback=ws.cell(row=x, column=11).value, _15daysCompletionRate=ws.cell(row=x, column=12).value, _20daysCompletionRate=ws.cell(row=x, column=13).value)
            _180GreaterPerformance = Performance(caseCount=ws.cell(row=x, column=14).value, amount=ws.cell(row=x, column=15).value, paybackAmount=ws.cell(row=x, column=16).value, paybackRate=ws.cell(row=x, column=17).value, targetRate=ws.cell(row=x, column=18).value, completionRate=ws.cell(
                row=x, column=19).value, completionAmount=ws.cell(row=x, column=20).value, order=ws.cell(row=x, column=21).value, _15daysPayback=ws.cell(row=x, column=22).value, _20daysPayback=ws.cell(row=x, column=23).value, _15daysCompletionRate=ws.cell(row=x, column=24).value, _20daysCompletionRate=ws.cell(row=x, column=25).value)
            overallPerformance = OverallPerformance(caseCount=ws.cell(row=x, column=26).value, amount=ws.cell(row=x, column=27).value, paybackAmount=ws.cell(row=x, column=28).value, paybackRate=ws.cell(row=x, column=29).value, targetOne=ws.cell(row=x, column=30).value, targetTwo=ws.cell(row=x, column=31).value, targetThree=ws.cell(
                row=x, column=32).value, targetThreeCompletionRate=ws.cell(row=x, column=33).value, targetOnePaybackAmount=ws.cell(row=x, column=34).value, targetTwoPaybackAmount=ws.cell(row=x, column=35).value, targetThreePaybackAmount=ws.cell(row=x, column=36).value, order=ws.cell(row=x, column=37).value)
            performanceReport = PerformanceReport(companyName=ws.cell(
                row=x, column=1).value, _180Performance=_180Performance, _180GreaterPerformance=_180GreaterPerformance, overallPerformance=overallPerformance)
            dataList.append(performanceReport)
    return dataList


def generateHTML(companyName, sortedByOverall, sortedBy180, sortedBy180Greater):
    THIS_DIR = os.path.dirname(os.path.abspath(__file__))
    j2_env = Environment(loader=FileSystemLoader(THIS_DIR), trim_blocks=True)
    h = j2_env.get_template('content.html').render(
        companyName=companyName, sortedByOverall=sortedByOverall, sortedBy180=sortedBy180, sortedBy180Greater=sortedBy180Greater)
    # print(h)
    return h


def main():
    # 获得数据
    dataList = loadRowData()
    # 排序
    sortedBy180 = sorted(dataList, key=lambda x: int(
        x._180Performance.order.split('/')[0]))[:_180_SIZE]

    sortedBy180Greater = sorted(dataList, key=lambda x: int(
        x._180GreaterPerformance.order.split('/')[0]))[:_180GREATER_SIZE]

    sortedByOverall = sorted(dataList, key=lambda x: int(
        x.overallPerformance.order.split('/')[0]))[:OVERALL_SIZE]
    # 登录邮件服务器
    smtp = smtplib.SMTP_SSL(HOST, port=465)
    smtp.login(SENDER, PASSWORD)
    # 执行每一个发送
    for data in dataList:
        sortedByOverall = sortedByOverall[:OVERALL_SIZE]
        sortedBy180 = sortedBy180[:_180_SIZE]
        sortedBy180Greater = sortedBy180Greater[:_180GREATER_SIZE]
        if int(data.overallPerformance.order.split('/')[0]) > OVERALL_SIZE:
            sortedByOverall.append(data)
        if int(data._180Performance.order.split('/')[0]) > _180_SIZE:
            sortedBy180.append(data)
        if int(data._180GreaterPerformance.order.split('/')[0]) > _180GREATER_SIZE:
            sortedBy180Greater.append(data)
        content = generateHTML(data.companyName, sortedByOverall,
                               sortedBy180, sortedBy180Greater)

        # 调用发送邮件
        contacts = loadContacts()
        generateEmailAndSend(data.companyName, contacts, content, smtp)
    # 关闭smtp对象
    smtp.quit


# 显示前几名的数据
OVERALL_SIZE = 5
_180_SIZE = 3
_180GREATER_SIZE = 3
# 配置发件人邮箱
HOST = 'smtp.xxxx.yy.com'
SENDER = 'xxxxxxx@xx.com'
PASSWORD = 'xxxxxx'

# 执行
main()
